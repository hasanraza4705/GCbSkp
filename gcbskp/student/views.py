from django.shortcuts import render
from gcbskp.student.models import Student
from django.contrib import messages
from django.shortcuts import  redirect
from django.shortcuts import get_object_or_404
from gcbskp.student.models import Application
from gcbskp.online_admission.models import AcademicRecord
from gcbskp.online_admission.models import Subject
from django.urls import reverse
from django.core.exceptions import ValidationError
from datetime import datetime
from django.core.paginator import Paginator
from django.http import HttpResponse
import csv
from gcbskp.offered_program.models import Programs
from gcbskp.coursegroup.models import CourseGroup
from gcbskp.student.merit_calculation import get_merit_calculator
from gcbskp.student_registration.models import RegisteredStudent
from django.http import JsonResponse
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import pdb


# add student data function
def add_student(request):
    if request.method == 'POST':
    
        name = request.POST.get('name')
        cnic_no = request.POST.get('cnic_no')
        photo = request.FILES.get('photo')
        date_of_birth = request.POST.get('date_of_birth')
        mobile_no = request.POST.get('mobile_no')
        std_email = request.POST.get('std_email')
        father_name = request.POST.get('father_name')
        father_cnic = request.POST.get('father_cnic')
        father_mobile_no = request.POST.get('father_mobile_no')
        guardian_name = request.POST.get('guardian_name')
        guardian_cnic = request.POST.get('guardian_cnic')
        guardian_contact_no = request.POST.get('guardian_contact_no')
        father_occupation = request.POST.get('father_occupation')
        permanent_address = request.POST.get('permanent_address')
        postal_address = request.POST.get('postal_address')
        province = request.POST.get('province')
        district = request.POST.get('district')
        city = request.POST.get('city')
        gender = request.POST.get('gender')
        religion = request.POST.get('religion')
        hafiz_e_quran = request.POST.get('hafiz_e_quran') == 'Yes' 
        hafiz_doc = request.FILES.get('hafiz_document')
        blood_group = request.POST.get('blood_group')
        marital_status = request.POST.get('marital_status')
        disability_status = request.POST.get('disabil_status') == 'Yes'
        disability_type = request.POST.get('disbility_type')
        related_to_worker = request.POST.get('related_to_worker') == 'Yes'  
        worker_name = request.POST.get('worker_name')
        worker_income = request.POST.get('worker_income')
        worker_status = request.POST.get('worker_status') == 'Yes'  
        payscale = request.POST.get('worker_payscale')

        user_id = request.session.get('user_id')
        User = get_object_or_404(RegisteredStudent, id=user_id)

        required_fields = {
            'name': name.title(),
            'cnic_no': cnic_no,
            'date_of_birth': date_of_birth,
            'mobile_no': mobile_no,
            'email': std_email,
            'father_name': father_name.title(),
            'father_cnic': father_cnic,
            'father_mobile_no': father_mobile_no,
            'father_occupation': father_occupation.title(),
            'permanent_address': permanent_address,
            'postal_address': postal_address,
            'province': province.title(),
            'district': district.title(),
            'city' : city.title(),
            'gender': gender,
            'religion': religion,
            'blood_group': blood_group,
            'marital_status': marital_status,
        }

        # Check for any empty required fields
        null_fields = [field for field, value in required_fields.items() if not value]

        if null_fields:
            null_field_names = ', '.join(null_fields)
            messages.error(request, f'The following fields are required: {null_field_names}')
            return render(request, 'online_admission:apply')

        try:
            # Check if a student with the given CNIC already exists
            if Student.objects.filter(cnic_no=cnic_no).exists():
                messages.error(request, 'A student with this CNIC already exists.')
                return render(request, 'online_admission:apply')
            if  User.cnic_no != cnic_no:
                messages.error(request, 'CNIC not not match with the registered user cnic.')
                return render(request, 'online_admission:apply')
            if  User.email != std_email:
                messages.error(request, 'Email not match with the registered user email.')
                return render(request, 'online_admission:apply')
            # Validate Image Size (Max 500KB)
            if photo:
                max_size = 500 * 1024  # 500KB
                if photo.size > max_size:
                    messages.error(request, "The uploaded image exceeds the maximum size of 500KB.")
                    return render(request, 'online_admission:apply')
            
            student = Student.objects.create(
                name=name,
                cnic_no=cnic_no,
                photo=photo,
                date_of_birth=date_of_birth,
                mobile_no=mobile_no,
                email=std_email,
                father_name=father_name,
                father_cnic=father_cnic,
                father_mobile_no=father_mobile_no,
                guardian_name=guardian_name,
                guardian_cnic=guardian_cnic,
                guardian_contact_no=guardian_contact_no,
                father_occupation=father_occupation,
                permanent_address=permanent_address,
                postal_address=postal_address,
                province=province,
                district=district,
                city=city,
                gender=gender,
                religion=religion,
                hafiz_e_quran=hafiz_e_quran,
                hafiz_doc = hafiz_doc,
                blood_group=blood_group,
                marital_status=marital_status,
                disability_status=disability_status,
                disability_type = disability_type,
                related_to_worker=related_to_worker,
                worker_name=worker_name if related_to_worker else None,
                worker_relation=worker_income if related_to_worker else None,
                worker_payscale=payscale if related_to_worker else None,
                worker_status = worker_status if related_to_worker else False,

                user_id=request.session.get('user_id')  
            )

            
            student.save()
            response = HttpResponse()
            request.session['student_application_id'] = student.id
            messages.success(request, 'Personal info stored successfully!')
            response.set_cookie('formSubmittedSuccessfully', 'true', max_age=3600)
            return redirect('online_admission:view_academic_record')
        except Exception as e:
            
            messages.error(request, f'Error submitting student application: {str(e)}')

        return render(request, 'online_admission:apply')
    
    return render(request, 'online_admission:apply')



def get_course_groups(request):
    program_id = request.GET.get('program_id')
    course_groups = CourseGroup.objects.filter(program_id=program_id)
    data = list(course_groups.values('id', 'name'))
    return JsonResponse(data, safe=False)

def calculate_deduction(passing_year):
    current_year = datetime.now().year  # Get the current year
    if passing_year:
        passing_year = int(passing_year)  # Ensure it's an integer
        years_since_passing = current_year - passing_year  # Calculate total years

        if years_since_passing > 2:
            deduction = (years_since_passing - 2) * 2  # (Total years - 2) * 2
            return deduction
    return 0  # No deduction if within 2 years


# merit calculation function
def generate_merit_list(request):
    if request.method == 'POST':
        program_id = request.POST.get('program_id')
        course_group_ids = request.POST.getlist('course_group_ids')

        # Retrieve the selected program
        program = Programs.objects.get(id=program_id)

        # Get the current year
        current_year = datetime.now().year

        # Initialize a workbook for multiple merit lists
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove the default sheet

        # Track students and their merit data for final update
        students_to_update = {}

        # Loop through each selected course group
        for course_group_id in course_group_ids:
            course_group = CourseGroup.objects.get(id=course_group_id)
            print(f"Processing Course Group: {course_group.name}")

            # Retrieve applications for the selected program and course group
            applications = Application.objects.filter(
                program=program,
                course_groups__id=course_group_id,
                admission_status='pending'
            )

            print(f"Number of Applications for {course_group.name}: {applications.count()}")

            # Get the merit calculator function based on the program
            merit_calculator = get_merit_calculator(program.heading, course_group.name)

            merit_list = []

            for app in applications:
                student = Student.objects.get(user_id=app.user_id)
                print(f"Student ID: {student.id}")
                print(f"Name: {student.name}")
                print(f"Father's Name: {student.father_name}")
                print(f"CNIC: {student.cnic_no}")
                print(f"Date of Birth: {student.date_of_birth}")

                # Retrieve FormData based on the application ID (not class_name)
                form_data = AcademicRecord.objects.filter(user_id=app.user_id)
                print(f"Form Data: {form_data}")

                if form_data.exists():
                    # Initialize variables for obtained marks and totals
                    matric_obtained = 0
                    matric_total = 1  # Avoid division by zero
                    inter_obtained = 0
                    inter_total = 1  # Avoid division by zero
                    matric_data = None
                    inter_data = None  # Initialize the inter_data variable

                    # Iterate through the form data and calculate merit based on available data
                    for data in form_data:
                        if data.class_name in ['Matric Science', 'Matric Arts', '0 Level']:
                            # Matriculation Data
                            matric_obtained = data.obtain_marks
                            matric_total = data.total_marks
                            matric_data = data  # Assign matric data
                            print(f"Matric Data: Obtained {matric_obtained}, Total {matric_total}")

                        elif data.class_name in ['ICS', 'FSC Pre Engineering', 'FSC Pre Medical', 'FA', 'FAIT', 'A Level']:
                            # Intermediate Data
                            inter_obtained = data.obtain_marks
                            inter_total = data.total_marks
                            inter_data = data  # Assign inter data
                            print(f"Intermediate Data: Obtained {inter_obtained}, Total {inter_total}")

                    # Calculate merit using the selected function
                    merit = merit_calculator(
                        matric_obtained=matric_obtained,
                        matric_total=matric_total,
                        inter_obtained=inter_obtained,
                        inter_total=inter_total,
                        hafiz_e_quran=student.hafiz_e_quran
                    )
                    print(f"Calculated Merit: {merit}")
                    if inter_data == 'None':
                        if matric_data and matric_data.year:
                            matric_deduction = calculate_deduction(matric_data.year)
                            merit -= matric_deduction
                        # Convert matric_data.year to an integer for comparison
                            # matric_year = int(matric_data.year)
                            # if matric_year < current_year - 2:  # Check if matric_year exists
                                # merit -= 2  # Reduce merit for older matric data
                            # elif matric_year != current_year:
                                # pass  # No adjustment needed for current year
                    else:
                        # Adjust merit based on the pass year
                        # Matric Pass Year Adjustment
                        if matric_data and matric_data.year:
                            matric_deduction = calculate_deduction(matric_data.year)
                            merit -= matric_deduction
                            # Convert matric_data.year to an integer for comparison
                            # matric_year = int(matric_data.year)
                            # if matric_year < current_year - 2:  # Check if matric_year exists
                                # merit -= 2  # Reduce merit for older matric data
                            # elif matric_year != current_year:
                                # pass  # No adjustment needed for current year


                    # Intermediate Pass Year Adjustment
                    if inter_data and inter_data.year:
                        inter_deduction = calculate_deduction(inter_data.year)
                        merit -= inter_deduction
                        # Convert inter_data.year to an integer for comparison
                        # inter_year = int(inter_data.year)
                        # if inter_year < current_year - 2:  # Check if inter_data exists
                            # pass  # No adjustment needed for old inter data
                        # elif inter_year != current_year:
                            # merit -= 2  # Reduce merit for older inter data

                    # Add to merit list
                    merit_list.append({
                        'student': student,
                        'merit': merit
                    })

                    # Store student merit data to update later
                    students_to_update[student.id] = {
                        'admission_status': 'admit',
                        'merit': merit
                    }

            print(f"Number of Merit List Entries for {course_group.name}: {len(merit_list)}")

            # Create a sheet for each course group
            ws = wb.create_sheet(title=f'{course_group.name} Merit List')

             # Define the column headers
            headers = [
                'Student Name',
                'Father\'s Name',
                'CNIC',
                'Date of Birth',
                'Mobile No',
                'Province',
                'Email',
                'Gender',
                'Class',
                'Year',
                'Board',
                'Matric Obtained Marks',
                'Matric Total Marks',
                'Intermediate Obtained Marks',
                'Intermediate Total Marks',
                'Class',
                'Year',
                'Board',
                'Merit'
            ]
            ws.append(headers)

            for item in merit_list:
                student = item['student']
                matric_data = AcademicRecord.objects.filter(user_id=student.user_id, class_name__in=['Matric Science', 'Matric Arts', '0 Level']).first()
                inter_data = AcademicRecord.objects.filter(user_id=student.user_id, class_name__in=['ICS', 'FSC Pre Engineering', 'FSC Pre Medical', 'FA', 'FAIT', 'A Level']).first()
                # Add the data to the Excel sheet
                ws.append([
                    student.name.title(),
                    student.father_name.title(),
                    student.cnic_no,
                    student.date_of_birth,
                    student.mobile_no,
                    student.province,
                    student.email,
                    student.gender,
                    matric_data.class_name,
                    matric_data.year,
                    matric_data.board,
                    matric_data.obtain_marks if matric_data else '',  # Matric obtained marks
                    matric_data.total_marks if matric_data else '',  # Matric total marks
                    inter_data.obtain_marks if inter_data else '',  # Intermediate obtained marks
                    inter_data.total_marks if inter_data else '',  # Intermediate total marks
                    inter_data.class_name,
                    inter_data.year,
                    inter_data.board,
                    item['merit']
                ])   
           
                  
            

            

            # Adjust column widths (optional)
            for col_num, col in enumerate(ws.columns, 1):
                max_length = max(len(str(cell.value)) for cell in col)
                ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

        # After processing all course groups, update the admission status of students
        for student_id, student_data in students_to_update.items():
            student = Student.objects.get(id=student_id)
            app = Application.objects.get(user_id=student.user_id, program=program)
            # Update the admission status and merit for the student
            if app.admission_status != 'admit':  # Ensure we only update if it's not already 'admit'
                app.admission_status = student_data['admission_status']
                app.save()

            # Optionally, you can also update any merit field in the student object, if necessary
            # student.merit = student_data['merit']
            # student.save()

        # Create the HttpResponse to send the Excel file to the user
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="multiple_merit_lists.xlsx"'

        # Save the workbook to the HttpResponse
        wb.save(response)

        return response

    return redirect('online_admission:dashboard')






def update_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        # Update student 
        student.name = request.POST.get('name', student.name)
        student.mobile_no = request.POST.get('std_modile_no', student.mobile_no)
        student.cnic_no = request.POST.get('std_cnic', student.cnic_no)
        # student.date_of_birth = request.POST.get('std_dob', student.date_of_birth)
         # Convert and update date_of_birth
        # dob_str = request.POST.get('std_dob', student.date_of_birth)
        # if dob_str:
        #     try:
        #         # Attempt to parse the date string into a datetime object
        #         parsed_date = datetime.strptime(dob_str, '%b. %d, %Y')
        #         student.date_of_birth = parsed_date.date()
        #     except ValueError:
        #         # Handle invalid date format
        #         student.date_of_birth = None
        
        student.father_name = request.POST.get('std_father_name', student.father_name)
        student.father_cnic = request.POST.get('father_cnic', student.father_cnic)
        student.father_mobile_no = request.POST.get('std_cnic', student.father_mobile_no)
        student.father_occupation = request.POST.get('father_occupation', student.father_occupation)
        student.guardian_name = request.POST.get('guardian_name', student.guardian_name)
        student.guardian_contact_no = request.POST.get('std_guardian_no', student.guardian_contact_no)
        student.guardian_cnic = request.POST.get('std_duardian_cnic', student.guardian_cnic)
        
        student.religion = request.POST.get('religion', student.religion)
        student.gender = request.POST.get('std_gender', student.gender)
        student.marital_status = request.POST.get('marital_status', student.marital_status)
        student.blood_group = request.POST.get('blood_group', student.blood_group)
        student.worker_relation = request.POST.get('worker_relation', student.worker_relation)
        student.worker_name = request.POST.get('worker_name', student.worker_name)
        student.permanent_address = request.POST.get('permanent_address', student.permanent_address)
        student.postal_address = request.POST.get('postal_address', student.postal_address)
        student.district = request.POST.get('district', student.district)

        # Parse boolean fields for Disability Status and Hafiz-e-Quran
        student.disability_status = request.POST.get('disability_status') == 'true'
        student.hafiz_e_quran = request.POST.get('hafiz_e_quran') == 'true'
        
        # Save the updated student record
        student.save()

        # Update form data
        form_data_ids = [key.split('_')[2] for key in request.POST.keys() if key.startswith('form_data_')]
        for form_data_id in form_data_ids:
            try:
                form_data = AcademicRecord.objects.get(id=form_data_id)
                form_data.class_name = request.POST.get(f'form_data_{form_data_id}_class_name', form_data.class_name)
                form_data.year = request.POST.get(f'form_data_{form_data_id}_year', form_data.year)
                form_data.board = request.POST.get(f'form_data_{form_data_id}_board', form_data.board)
                form_data.obtain_marks = request.POST.get(f'form_data_{form_data_id}_obtain_marks', form_data.obtain_marks)
                form_data.percentage = request.POST.get(f'form_data_{form_data_id}_percentage', form_data.percentage)
                form_data.save()
            except AcademicRecord.DoesNotExist:
                continue

        # Update subjects
        subject_ids = [key.split('_')[1] for key in request.POST.keys() if key.startswith('subject_')]
        for subject_id in subject_ids:
            try:
                subject = Subject.objects.get(id=subject_id)
                subject.subject_name = request.POST.get(f'subject_{subject_id}_name', subject.subject_name)
                subject.obtain_marks = request.POST.get(f'subject_{subject_id}_obtain_marks', subject.obtain_marks)
                subject.total_marks = request.POST.get(f'subject_{subject_id}_total_marks', subject.total_marks)
                subject.save()
            except Subject.DoesNotExist:
                continue
        messages.success(request, 'Student  updated  successfully!')
        return redirect('online_admission:studentList')  # Redirect to the student list view

    return redirect('online_admission:studentList')  # Redirect if not POST request


# student promoted process view
def update_verify_status(request):
    if request.method == 'POST':
        student_ids = request.POST.getlist('students')
        if student_ids:
            Student.objects.filter(id__in=student_ids).update(verifiy_status='promoted')
            messages.success(request, 'Selected students have been promoted.')
        else:
            messages.warning(request, 'No students selected.')
    return redirect('online_admission:studentList') 

# promoted student get view
def promotedStudent(request):
    students = Student.objects.filter(verifiy_status='promoted')  # or another queryset based on your requirement
    student_records = []

    for student in students:
        form_data_list = AcademicRecord.objects.filter(user_id=student.user_id)
        student_subjects = []

        for form_data in form_data_list:
            subjects_for_data = Subject.objects.filter(form_data_id=form_data.id)
            student_subjects.append({
                'form_data': form_data,
                'subjects': subjects_for_data
            })

        application_records = []
        applications = Application.objects.filter(user_id=student.user_id)
        for application in applications:
            related_course_groups = application.course_groups.all()
            application_records.append({
                'application': application,
                'course_groups': related_course_groups
            })

        student_records.append({
            'student': student,
            'form_data': form_data_list,
            'form_data_subjects': student_subjects,
            'applications': application_records,
        })

    paginator = Paginator(student_records, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    program= Programs.objects.all()
    Coursegroup = CourseGroup.objects.all()

    context = {
        'student_records': page_obj,
        'page_obj': page_obj,
        'programs':program,
        'course_groups':Coursegroup
    }

    return render(request, 'pages/promotedStudentList.html', context)


# Add academic record data
def AddData(request):
    response = HttpResponse()  # Initialize response here
    application_id = request.session.get('student_application_id')
    if application_id is None:
        print(f"Number of Applicationsjk for {application_id}")
        return redirect('home:sliderData')
    if request.method == 'POST':
        class_name =  (request.POST.get('matric_stream') or '') or (request.POST.get('generic_class_name') or '') or (request.POST.get('name') or '').title()
        year = request.POST.get('year')
        board = request.POST.get('board')
        obtain_marks = request.POST.get('obtain_marks')
        total_marks = request.POST.get('total_marks')
        grade = request.POST.get('grade') or request.POST.get('cgpa')
        percentage = request.POST.get('percentage')

        # Validate and convert marks to integers
        try:
            obtain_marks = int(obtain_marks)
            total_marks = int(total_marks)
           
        except ValueError:
            messages.error(request, 'Obtained marks, total marks, and percentage must be valid numbers.')
            return redirect(reverse('online_admission:view_academic_record'))

        if total_marks < obtain_marks:
            messages.error(request, 'Total marks cannot be less than obtained marks.')
            return redirect(reverse('online_admission:view_academic_record'))

        if not application_id:
            messages.error(request, 'Data not submitted successfully because application not submitted.')
            return redirect(reverse('online_admission:view_academic_record'))
        if AcademicRecord.objects.filter(class_name=class_name,application_id=application_id).exists():
            messages.error(request, f'The class "{class_name}" has already been added for this academic record.')
            return redirect(reverse('online_admission:view_academic_record'))

        try:
            formData = AcademicRecord(
                class_name=class_name,
                year=year,
                board=board,
                obtain_marks=obtain_marks,
                total_marks=total_marks,
                grade=grade,
                percentage=percentage,
                user_id=request.session.get('user_id'),
                application_id=application_id,
            )
            formData.save()
            request.session['form_data_id'] = formData.id
            messages.success(request, 'Data added successfully.')
        except Exception as e:
            messages.error(request, f'Error adding data: {str(e)}')
            return redirect(reverse('online_admission:view_academic_record'))

        return redirect(reverse('online_admission:subject_view'))

    # Handle the case where the method is not POST
    messages.error(request, 'Invalid request method.')
    return redirect(reverse('online_admission:view_academic_record'))
# Add subject

def add_subject(request):
    response = HttpResponse()  # Initialize response at the start

    if request.method == 'POST':
        # Get the form data
        academic_record_id = request.POST.get('academic_record_id')  # Get the selected academic record ID
        subject_name = request.POST.get('subject_name').title()
        subject_obtaining_marks = request.POST.get('subject_obtaining_marks')
        subject_total_marks = request.POST.get('subject_total_marks')

        # Convert marks to integers
        try:
            subject_obtaining_marks = int(subject_obtaining_marks)
            subject_total_marks = int(subject_total_marks)
        except ValueError:
            messages.error(request, 'Obtaining marks and total marks must be valid numbers.')
            response.set_cookie('formSubmittedSuccessfully', 'true', max_age=3600)
            return redirect(reverse('online_admission:subject_view'))

        if subject_total_marks < subject_obtaining_marks:
            messages.error(request, 'Total marks cannot be less than obtained marks.')
            response.set_cookie('formSubmittedSuccessfully', 'true', max_age=3600)
            return redirect(reverse('online_admission:subject_view'))

        # Ensure academic_record_id is selected
        if not academic_record_id:
            messages.error(request, 'Please select an academic record.')
            response.set_cookie('formSubmittedSuccessfully', 'true', max_age=3600)
            return redirect(reverse('online_admission:subject_view'))
        
        if Subject.objects.filter(subject_name=subject_name,form_data_id=academic_record_id).exists():
            messages.error(request, f'The same subjects has already been added for this academic record.')
            return redirect(reverse('online_admission:subject_view'))

        try:
            # Create the subject
            Subject.objects.create(
                subject_name=subject_name,
                obtain_marks=subject_obtaining_marks,
                total_marks=subject_total_marks,
                form_data_id=academic_record_id,  # Store the selected academic record ID
                user_id=request.session.get('user_id')  # Assuming user ID is stored in the session
            )
            messages.success(request, 'Subject added successfully!')
            response.set_cookie('formSubmittedSuccessfully', 'true', max_age=3600)
        except Exception as e:
            messages.error(request, f'Error adding subject: {str(e)}')
            response.set_cookie('formSubmittedSuccessfully', 'true', max_age=3600)

        return redirect(reverse('online_admission:subject_view'))

    # If not POST request, redirect with an error message
    messages.error(request, 'Invalid request method.')
    response.set_cookie('formSubmittedSuccessfully', 'true', max_age=3600)
    return redirect(reverse('online_admission:subject_view'))

def delete_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    subject.delete()
    return redirect('online_admission:subject_view')

def print_challan(request):
    user_id = request.session.get('user_id')
    student = get_object_or_404(Student, user_id=user_id)
    form_data = AcademicRecord.objects.filter(user_id=user_id)
    subjects = Subject.objects.filter(user_id=user_id)
    application = get_object_or_404(Application, user_id=user_id)

    context = {
        'student': student,
        'form_data': form_data,
        'subjects': subjects,
        'application': application,
        'user_id':user_id
    }
    return render(request, 'pages/fee_challan.html', context)